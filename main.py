import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox

# Функции для выполнения программ
def program1(convolution.py):
    print(f"Выполняется Program 1 с файлом: {convolution.py}")
    import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re

def extract_type_from_brackets(value_b):
    matches = re.findall(r'\((.*?)\)', value_b)
    if matches:
        return matches[0]
    else:
        return ""
first_row_skipped = True

def create_new_excel_file(input_file_path, output_file_path):
    workbook = openpyxl.load_workbook(input_file_path)
    
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    
    header = ["Cabinet", "Location", "Type", "Channel", "KKS", "CIR", "CONNECTION"]
    new_sheet.append(header)
    
    buff_a = None  
    next_buff_a = None  
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        i = 1
        j = 2
        
        while True:
            buff_a = None
            next_buff_a = None
            processed_rows = False
            
            for row in sheet.iter_rows(min_row=1, max_col=52, max_row=sheet.max_row):
                value_a = row[0].value
                value_b = row[i].value
                
                if isinstance(value_a, str) and value_a.isalpha():
                    if isinstance(value_b, str):
                        type_value = extract_type_from_brackets(value_b)
                        value_b = re.sub(r'\([^)]*\)', '', value_b)
                        location_value = value_a + value_b.lstrip('0')
                        if buff_a is None:
                            buff_a = row[0].row  
                        next_buff_a = row[0].row + 1 
                        first_row_skipped = False
                        

                if buff_a is None:
                    continue  

               
                if not first_row_skipped:
                  first_row_skipped = True
                  continue

                if row[i].fill.start_color.index != '00000000':  
                    channel_value = int(value_a)
                    kks = str(value_b)
                    cir = str(row[j].value)
                    new_sheet.append([sheet_name, location_value, type_value, channel_value, kks, cir, ''])
                    processed_rows = True


            i += 2
            j += 2
            if i > 52 or j > 52:  
                 break

            # all_white = True
            # for row in sheet.iter_rows(min_row=1, max_col=52, max_row=sheet.max_row):
            #     if i <= 52 and row[i].fill.start_color.index != '00000000':
            #         all_white = False
            #         break
            #     if j <= 52 and row[j].fill.start_color.index != '00000000':
            #         all_white = False
            #         break
            
            # if all_white:
            #     break 

            

    
    new_sheet.auto_filter.ref = new_sheet.dimensions  
    new_sheet.auto_filter.add_sort_condition("A2:B{}".format(new_sheet.max_row)) 

    new_workbook.save(output_file_path)
    print(f"Новый файл Excel сохранен как: {output_file_path}")

input_file_path = '/content/drive/MyDrive/Свертка_таблицы/10CRC11.xlsx'
output_file_path = '/content/drive/MyDrive/Свертка_таблицы/test.xlsx'

create_new_excel_file(input_file_path, output_file_path)

def program2(sort.py):
    print(f"Выполняется Program 2 с файлом: {sort.py}")
    import pandas as pd
import re

# Функция для извлечения префикса и числовой части из строки
def split_string(s):
    match = re.match(r"([A-Za-z]+)(\d+)", s)
    if match:
        return match.groups()[0], int(match.groups()[1])
    else:
        return s, -1  # Возвращаем исходную строку и -1 для строк, не соответствующих шаблону

# Функция для сортировки с использованием split_string
def custom_sort(val):
    prefix, number = split_string(val)
    return (prefix, number)

# Открытие файла Excel
file_path = "/content/drive/MyDrive/Свертка_таблицы/test.xlsx"
df = pd.read_excel(file_path)

# Сортировка по двум столбцам с кастомной функцией сортировки для каждого столбца
sorted_df = df.sort_values(by=[df.columns[0], df.columns[1]], key=lambda col: col.map(custom_sort))

# Сохранение результата в новый файл
sorted_df.to_excel("/content/drive/MyDrive/Свертка_таблицы/sorted_test.xlsx", index=False)

# Функции для выбора файлов
def choose_file1():
    file_path1 = filedialog.askopenfilename()
    file_label1.config(text=file_path1)

def choose_file2():
    file_path2 = filedialog.askopenfilename()
    file_label2.config(text=file_path2)

# Функция для выполнения выбранных программ
def execute_programs():
    file_path1 = file_label1.cget("text")
    file_path2 = file_label2.cget("text")

    if checkbox_var1.get() and not file_path1:
        messagebox.showwarning("Предупреждение", "Пожалуйста, выберите файл для Program 1.")
        return

    if checkbox_var2.get() and not file_path2:
        messagebox.showwarning("Предупреждение", "Пожалуйста, выберите файл для Program 2.")
        return
    
    if checkbox_var1.get():
        program1(file_path1)
    if checkbox_var2.get():
        program2(file_path2)

# Создание главного окна
root = tk.Tk()
root.title("Программа с визуализацией")

# Переменные для чекбоксов
checkbox_var1 = tk.BooleanVar()
checkbox_var2 = tk.BooleanVar()

# Чекбоксы для выбора программ
checkbox1 = tk.Checkbutton(root, text="Program 1", variable=checkbox_var1)
checkbox1.pack()

# Метка и кнопка для выбора файла для Program 1
file_label1 = tk.Label(root, text="Файл для Program 1 не выбран")
file_label1.pack()
choose_button1 = tk.Button(root, text="Выбрать файл для Program 1", command=choose_file1)
choose_button1.pack()

checkbox2 = tk.Checkbutton(root, text="Program 2", variable=checkbox_var2)
checkbox2.pack()

# Метка и кнопка для выбора файла для Program 2
file_label2 = tk.Label(root, text="Файл для Program 2 не выбран")
file_label2.pack()
choose_button2 = tk.Button(root, text="Выбрать файл для Program 2", command=choose_file2)
choose_button2.pack()

# Кнопка для выполнения выбранных программ
execute_button = tk.Button(root, text="Выполнить", command=execute_programs)
execute_button.pack()
